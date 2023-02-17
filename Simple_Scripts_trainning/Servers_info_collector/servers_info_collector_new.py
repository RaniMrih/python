
#!/usr/bin/env python
import os
import sys
import subprocess
import argparse

from Common import *
from Config import *
from DataBaseClass import *

logger.set_level(LevelEnum.INFO)

import warnings
warnings.simplefilter("ignore")
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from openpyxl.styles import PatternFill, Border, Side, Alignment

def get_servers():
    servers_str = "l-fwbld-[01-04],l-fwreg-[001-142,150-157,160-167,250],l-fwminireg-[010-014,020-024,030-034,040-044,050-054,060-064,070-074,080-084,090-094],l-fwdev-[002-012,014-175],l-fwvrt-[01,06,09],l-fwbld-clx-[01-08,13]"
    #servers_str = "l-fwbld-[01-04],l-fwreg-[001-147,150-157,160-167,250],l-fwminireg-[010-014,020-024,030-034,040-044,050-054,060-064,070-074,080-084,090-094]"
    # servers_str = "l-fwreg-[001-008,010-038,041-051,053-112]"
    # servers_str = "l-fwdev-[129-140],l-fwreg-[009,039,040,052,228-231]"
    servers_str = "l-fwreg-[001-149]"
    logger.info("servers str ==> %s" % servers_str)
    servers_arr = []
    server_prefix = ""
    ingore_indexes = -1
    for index, server_char in enumerate(servers_str):
        if ( ingore_indexes >= index ):
            continue
        if ( server_char == "," ):
            if ( server_prefix != ""  ):
                servers_arr.append(server_prefix)
            server_prefix = ""
            continue
        if ( server_char != "[" and server_char != "]" ):
            server_prefix += server_char
        if ( server_char == "[" ):
            start_index = index +1
            ignore_range_indexs = -1
            end_index = servers_str.index("]", start_index)
            range_str = ""
            for index_range_char in range(start_index, end_index):
                if ( ignore_range_indexs >= index_range_char ):
                    continue
                range_char = servers_str[index_range_char]
                
                if ( range_char != "," and range_char != "-" ):
                    range_str += range_char
                if ( range_char == "," ):
                    if (range_str!= ""):
                        servers_arr.append("%s%s"%(server_prefix, range_str))
                    range_str = ""
                if ( range_char == "-" ):
                    start = int(range_str)
                    end_range = servers_str[index_range_char+1]
                    end_rande_str = ""
                    while ( end_range >= "0" and end_range <= "9" ):
                        end_rande_str += end_range
                        index_range_char += 1
                        end_range = servers_str[index_range_char+1]
                    ignore_range_indexs = index_range_char
                    end = int(end_rande_str)
                    for i in xrange(start, end+1):
                        len_range= len(end_rande_str)
                        servers_arr.append("%s%s"%(server_prefix, str(i).zfill(len_range)))
                    range_str = ""
                    
            if ( range_str ):
                servers_arr.append("%s%s"%(server_prefix, range_str))
            ingore_indexes = end_index-1
            server_prefix = ""
            continue
    if ( server_prefix != "" ):
        servers_arr.append(server_prefix)
    logger.info("servers list ==> %s" % servers_arr)
    return servers_arr

def write_to_excel(db_connection):
    logger.info("write_to_excel")
    row_index = 1
    warnings.simplefilter("ignore")
    work_book = openpyxl.Workbook()
    
    work_sheet= work_book.create_sheet("Servers Data", 0)

    db_connection.row_factory = sqlite3.Row
    db_cursor = db_connection.cursor()
    for row in db_cursor.execute('SELECT * FROM ServersInfo ORDER BY NAME'):
        if ( row_index == 1 ):
            for index, column_name in enumerate(row.keys(), start=1):
                work_sheet.cell(column=index, row=row_index, value=column_name)
            row_index+= 1   
        for index,member in enumerate(row, start=1):
            work_sheet.cell(column=index, row=row_index, value=member)
        row_index += 1
    
    maxcolumnletter = openpyxl.utils.get_column_letter(work_sheet.max_column)
    ref = "A1:{max_col}{max_row}".format(max_col=maxcolumnletter, max_row=str(len(work_sheet['A'])) )
    tab = Table(displayName="Table1", ref=ref)

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    work_sheet.add_table(tab)
    
    work_sheet= work_book.create_sheet("Devices Data", 1)
    db_cursor = db_connection.cursor()
    row_index = 1
    db_connection.row_factory = sqlite3.Row

    for row in db_cursor.execute('SELECT * FROM DevicesInfo ORDER BY NAME'):
        if ( row_index == 1 ):
            for index, column_name in enumerate(row.keys(), start=1):
                work_sheet.cell(column=index, row=row_index, value=column_name)
            row_index+= 1  
        for index,member in enumerate(row, start=1):
            work_sheet.cell(column=index, row=row_index, value=member)
        row_index += 1

    maxcolumnletter = openpyxl.utils.get_column_letter(work_sheet.max_column)
    logger.debug("MAX ==> %s" % maxcolumnletter)
    tab = Table(displayName="Table1", ref="A1:"+maxcolumnletter+str(len(work_sheet['A'])))
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    
    
    ref = "A1:{max_col}{max_row}".format(max_col=maxcolumnletter, max_row=str(len(work_sheet['A'])) )
    tab = Table(displayName="Table2", ref=ref)

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    work_sheet.add_table(tab)
    
    work_book.save(filename = ExcelPath) 

def write_to_excel_recovery(db_connection):
    logger.info("write_to_excel")
    row_index = 1
    warnings.simplefilter("ignore")
    work_book = openpyxl.Workbook()
    
    # work_sheet= work_book.create_sheet("Servers Data", 0)
    work_sheet = work_book.active
    work_sheet.title = "Servers Data"

    db_connection.row_factory = sqlite3.Row
    db_cursor = db_connection.cursor()
    for row in db_cursor.execute('select ServersInfo.Name,DevicesInfo.Device_name,DevicesInfo.card_name,DevicesInfo.PCI_slot,ServersInfo.has_lf, ServersInfo.allocated, ServersInfo.status from ServersInfo INNER JOIN DevicesInfo ON ServersInfo.Name = DevicesInfo.Name ORDER BY ServersInfo.Name;'):
        if ( row_index == 1 ):
            for index, column_name in enumerate(row.keys(), start=1):
                if column_name == "Name": column_name="server_name"
                if column_name == "Device_name": column_name = "device_name"
                if column_name == "card_name": column_name = "cards"
                if column_name == "PCI_slot" : column_name = "pci"

                work_sheet.cell(column=index, row=row_index, value=column_name)
            row_index+= 1   
        for index,member in enumerate(row, start=1):
            if member == "True": member = "yes"
            if member == "False" or member == "": member = "no"
            work_sheet.cell(column=index, row=row_index, value=member)
        row_index += 1
    
    work_book.save(filename = ExcelPathRecovery) 


def parse_arguments (args):
    list_params = ["servers", "log_file"]
    
    parser = argparse.ArgumentParser(description='')
    
    parser.add_argument("-l",  "--servers", required=True, action='store', nargs=1, help="servers list")
    parser.add_argument("--debug", action='store_true', help="debug messages")
    parser.add_argument("--debug2", action='store_true', help="debug messages for sub-tool")
    parser.add_argument("--log_file", action='store', nargs=1, help="log file tp store the output")
    
    args_array = parser.parse_args(args[1:])
    for param in list_params:
        if (args_array.__getattribute__(param)):
            args_array.__setattr__(param, args_array.__getattribute__(param)[0])
    
    return args_array

if __name__ == "__main__":
    args_array = parse_arguments(sys.argv)
    if ( args_array.log_file ):
        logger = FwLogger("Get server info", args_array.log_file)
        logger.set_level()
    if ( args_array.debug ):
        logger.set_level(LevelEnum.DEBUG)

    if ( args_array.servers == "ALL" ):
        servers_list = get_servers()
    else:
        servers_list = args_array.servers.split(",")
    
    logger.debug("atefs")
    # exit(0)

    processes_list = list()
    sub_tool_debug = ""
    if ( args_array.debug2 ):
        sub_tool_debug = "--debug"
    
    for index in xrange(0, len(servers_list), 5):
        for index_2 in xrange(index,index+5):
            if index_2 >= len(servers_list): continue
            server = servers_list[index_2]
            logger.info("server name: %s" % server)
            server_cmd = "pdsh -w {server_name} 'python2 /.autodirect/fwgwork/atefs/workspace/scripts/get_servers_info/new_servers_colloctor/get_server_info.py --server_name {server_name} {debug}'".\
                    format(server_name=server, debug=sub_tool_debug)
            # server_cmd = "sshpass -p 'atefs' ssh atefs@{server_name} python2 /.autodirect/fwgwork/atefs/workspace/scripts/get_servers_info/new_servers_colloctor/get_server_info.py --server_name {server_name} {debug}" .\
            #     format(server_name=server, debug=sub_tool_debug)
            # server_cmd = "ssh {server_name} 'python /.autodirect/fwgwork/atefs/workspace/scripts/get_servers_info/new_servers_colloctor/get_server_info.py --server_name {server_name}'".\
            #         format(server_name=server)
            server_process = run_command (server_cmd, stderr_flag = False, wait_for_process=False)
            processes_list.append(server_process)
        import time
        time.sleep(5)
    
    final_rc = RC_SUCCESS
    for index,process in enumerate(processes_list):
        server_name = servers_list[index]
        result,err = process.communicate()
        result = result[0:len(result)-1]
        rc = process.returncode
        logger.debug("server {server} result: {res}".format(server=server_name, res=result))
        logger.info("server {server} rc: {rc}".format(server=server_name, rc=rc))
        final_rc = final_rc or rc
        if ( err ):
            logger.error("server {server} error: {err}".format(server=server_name, err=err))
            final_rc = final_rc or 1

    logger.info("final rc: %s" % final_rc)
    db_connection = sqlite3.connect(DataBasePath, check_same_thread=False)
    write_to_excel(db_connection)
    write_to_excel_recovery(db_connection)

    logger.exit_on_rc(final_rc)
