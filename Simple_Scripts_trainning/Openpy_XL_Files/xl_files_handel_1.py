import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    
    #wb shorten for workbook
    wb =xl.load_workbook(filename)
    sheet =  wb['Sheet1']

    for row in range (2, sheet.max_row + 1):
        # iterate on all rows taking cell 3
        cell = sheet.cell(row,3)
        corrected_price = cell.value * 0.9
        # writing cooreceted value to cell 4
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value = corrected_price

    #creating the graph obj
    values = Reference(sheet,
            min_row=2,
            max_row=sheet.max_row,
            min_col=4,
            max_col=4)

    #creating the chart
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,'e2')

    #overriding same file
    wb.save(filename)

if __name__ == '__main__':
   process_workbook('transactions.xlsx')
