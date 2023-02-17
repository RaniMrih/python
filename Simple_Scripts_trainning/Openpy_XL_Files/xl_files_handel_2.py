from openpyxl import Workbook, load_workbook
import sys
sys.path.append('c:/Users/rmrih/Rani_GitHub/python/Simple_Scripts_trainning/')
import general_classes as G

#------------------------------------basics---------------------
wb_path = G.Pathes.grades_wb_xl
wb = load_workbook(wb_path)

#active if there is only one sheet
ws = wb.active
print(ws)

#if want to choose sheet name
ws = wb['Sheet1']

#printing 1 cell
print(ws['A1'].value)

#changing cell value
ws['A2'].value= 'Test'
wb.save(wb_path)

print(wb.sheetnames)

#------------------------------------creating new sheet---------------------


